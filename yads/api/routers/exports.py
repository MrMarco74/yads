import logging
import json
import os
import shutil
import csv
from io import StringIO, BytesIO
from typing import Optional, List
from fastapi import APIRouter, Depends, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, RedirectResponse
from sqlmodel import Session, select, func, text
from datetime import datetime

from yads.database import get_session
from yads.auth.deps import RoleChecker
from yads.models import User, Target, ScanResult
from yads.api.templating import templates
from yads.core.backup import create_backup_zip, restore_backup_from_zip
from fpdf import FPDF

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("/api/backup/export")
async def export_data(
    tenant_ids: Optional[str] = Form(None),
    password: Optional[str] = Form(None),
    session: Session = Depends(get_session)
):
    """
    Generates and downloads a full or partial system backup (Zip).
    """
    try:
        # Parse tenant_ids
        t_ids_list = []
        if tenant_ids:
             try:
                 t_ids_list = [int(x.strip()) for x in tenant_ids.split(",") if x.strip()]
             except ValueError:
                 logger.warning(f"Invalid tenant_ids format: {tenant_ids}")
                 pass

        zip_file = create_backup_zip(session, tenant_ids=t_ids_list, password=password)
        
        # Determine extension based on encryption
        ext = "enc" if password else "zip"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"yads_backup_{timestamp}.{ext}"
        
        return StreamingResponse(
            zip_file, 
            media_type="application/octet-stream", # Generic binary
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/backup/analyze", response_class=HTMLResponse)
async def analyze_backup(
    request: Request,
    file: UploadFile = File(...),
    password: Optional[str] = Form(None),
    user: User = Depends(RoleChecker(["admin", "tenant_admin"])),
    session: Session = Depends(get_session)
):
    """
    Analyzes the uploaded backup file and returns a summary for confirmation.
    """
    contents = await file.read()
    
    meta = {}
    db_summary = {}
    
    # Try to handle potential encryption
    file_bytes = io.BytesIO(contents)
    is_encrypted = False
    
    # Attempt to open as Zip
    try:
        zf_check = zipfile.ZipFile(file_bytes, 'r')
        zf_check.close()
        file_bytes.seek(0)
    except zipfile.BadZipFile:
        # Might be encrypted
        if password:
            try:
                from yads.core.backup import decrypt_data
                decrypted = decrypt_data(contents, password)
                file_bytes = io.BytesIO(decrypted)
                is_encrypted = True
            except Exception as e:
                return HTMLResponse(f"<div class='text-red-400'>Decryption failed: {str(e)}</div>", status_code=400)
        else:
             return HTMLResponse("<div class='text-red-400'>Invalid Zip File. If encrypted, please provide password.</div>", status_code=400)

    try:
        with zipfile.ZipFile(file_bytes, 'r') as zf:
            if "metadata.json" in zf.namelist():
                meta = json.loads(zf.read("metadata.json"))
            
            # Count records roughly
            for name in zf.namelist():
                if name.startswith("data/") and name.endswith(".json"):
                    table = name.replace("data/", "").replace(".json", "")
                    data = json.loads(zf.read(name))
                    db_summary[table] = len(data)
    except Exception as e:
        logger.error(f"Error analyzing backup content: {e}")
        return HTMLResponse(f"<div class='text-red-400'>Error analyzing backup: {str(e)}</div>", status_code=400)

    # Encode contents to pass to next step? NO. Too large.
    # We should save to a temp file and confirm via ID?
    # Security risk: Temp file handling.
    # Or: The User re-uploads for confirmation (simpler stateless)?
    # OR: We use a signed token/cache.
    
    # SIMPLE APPROACH: Save to /tmp/yads_restore_pending.zip
    # Not thread safe for multiple admins restoring same time, but acceptable for this scope.
    import os
    tmp_path = "/tmp/yads_restore_pending.zip"
    with open(tmp_path, "wb") as f:
        f.write(contents)
    logger.info(f"Backup saved to temporary path: {tmp_path}")
        
    # Look up Tenant Names (Pre-load from DB first, then try to fill from Zip if unknown)
    tenant_ids = meta.get("tenant_ids", [])
    tenant_map = {tid: f"Unknown ID {tid}" for tid in tenant_ids}

    # 1. Try to read from Zip (Best source for restore)
    try:
        with zipfile.ZipFile(file_bytes, 'r') as zf:
            if "data/tenant.json" in zf.namelist():
                t_data = json.loads(zf.read("data/tenant.json"))
                for t in t_data:
                    if t.get("id") in tenant_map:
                        tenant_map[t.get("id")] = t.get("name")
    except Exception as e:
        logger.warning(f"Could not read tenant names from zip: {e}")

    # 2. Flatten for template
    tenant_names = [tenant_map.get(tid, f"ID {tid}") for tid in tenant_ids]
    
    from yads.core.backup import SYSTEM_TABLES
    # Render Confirmation Modal
    return templates.TemplateResponse("components/restore_confirmation_modal.html", {
        "request": request,
        "meta": meta,
        "db_summary": db_summary, # Assuming db_summary is now 'stats' in the new template context
        "tenant_names": tenant_names,
        "is_partial": bool(tenant_ids),
        "tmp_path": tmp_path,
        "skipped_tables": SYSTEM_TABLES,
        "filename": file.filename, # Added filename
        "password": password # Pass back to be embedded in hidden field
    })

@router.post("/api/backup/execute_restore")
async def execute_restore(
    confirmed: bool = Form(...),
    session: Session = Depends(get_session),
    user: User = Depends(RoleChecker(["admin"]))
):
    """
    Actually executes the restore from the temp file.
    """
    logger.info(f"Received restore execution request. Confirmed: {confirmed}")
    if not confirmed:
         return RedirectResponse(url="/settings?msg=Restore+Cancelled", status_code=303)
         
    import os
    tmp_path = "/tmp/yads_restore_pending.zip"
    if not os.path.exists(tmp_path):
        return RedirectResponse(url="/settings?error=Restore+Timeout:+File+not+found.+Please+upload+again.", status_code=303)
        
    try:
        with open(tmp_path, "rb") as f:
            content = f.read()
            
        from yads.core.backup import restore_backup_from_zip
        # Re-read meta for safety (logic inside restore anyway)
        restore_backup_from_zip(session, content)
        
        # Cleanup
        os.remove(tmp_path)
        
        return RedirectResponse(url="/settings?msg=System+Restored+Successfully.+Tenant+data+has+been+updated.", status_code=303)
    except Exception as e:
        logger.error(f"Restore Error: {e}")
        return RedirectResponse(url=f"/settings?error=Restore+Failed:+{str(e)}", status_code=303)
@router.get("/targets/export/excel")
async def export_targets_excel(lang: str = "en", session: Session = Depends(get_session), user: User = Depends(RoleChecker(["admin", "scanner"]))):
    """
    Generates an Excel report of all targets and their latest scan results.
    """
    import pandas as pd
    from io import BytesIO
    
    # Fetch targets for this tenant
    targets = session.exec(select(Target).where(Target.tenant_id == user.tenant_id).order_by(Target.created_at.desc())).all()
    
    # -- Cipher Compliance Setup --
    from yads.models import SystemConfig
    approved_ciphers_set = set()
    ac_conf = session.get(SystemConfig, "APPROVED_CIPHERS")
    raw_ac = ""
    if ac_conf:
        raw_ac = ac_conf.value
    else:
        # Load from file default
        try:
             import os
             if os.path.exists("ciphers.csv"):
                with open("ciphers.csv", "r") as f:
                    raw_ac = f.read()
        except Exception as e:
            logger.debug(f"Failed to load ciphers.csv fallback: {e}")
    
    if raw_ac:
        for line in raw_ac.splitlines():
             parts = line.split(',')
             if len(parts) >= 2:
                cipher_name = parts[1].strip()
                # Skip header if it exists
                if cipher_name and cipher_name.lower() != "cipherset":
                    approved_ciphers_set.add(cipher_name)
    
    data = []
    for t in targets:
        # Fetch latest results for each module
        results = session.exec(select(ScanResult).where(ScanResult.target_id == t.id).order_by(ScanResult.scanned_at.desc())).all()
        
        # Identify specific module results
        sub_scan = next((r for r in results if r.module_name == 'subdomain_scanner'), None)
        dns_scan = next((r for r in results if r.module_name == 'dns_scanner'), None)
        dns = sub_scan if sub_scan else dns_scan
        
        ssl = next((r for r in results if r.module_name == 'ssl_scanner'), None)
        web = next((r for r in results if r.module_name == 'web_analyzer'), None)
        infra = next((r for r in results if r.module_name == 'infrastructure_scanner'), None)
        tld_scan = next((r for r in results if r.module_name == 'tld_scanner'), None)
        port_scan = next((r for r in results if r.module_name == 'port_scanner'), None)
        
        # Online Status Logic
        is_online = "Unknown"
        if infra or web or port_scan:
            has_ip = bool(infra and infra.data and infra.data.get("ip"))
            has_http = False
            if web and web.data and web.data.get("status_code"):
                code = web.data.get("status_code")
                if isinstance(code, int) and code > 0:
                    has_http = True
            has_probe = bool(port_scan and port_scan.data and port_scan.data.get("is_active"))
            is_online = "Online" if (has_ip or has_http or has_probe) else "Offline"

        # DNS 
        dns_ip = ""
        if dns and dns.data and "records" in dns.data and "A" in dns.data["records"]:
             a_records = dns.data["records"]["A"]
             if a_records:
                 dns_ip = a_records[0]

        # Compliance
        compliant = 0
        non_compliant = 0
        if ssl and ssl.data and not ssl.data.get("error"):
            detected_ciphers = ssl.data.get("ciphers", [])
            for dc in detected_ciphers:
                name = dc.get("name")
                if name:
                    if name in approved_ciphers_set: compliant += 1
                    else: non_compliant += 1
        
        # CVEs
        cve_stats = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        if web and web.data and "cves" in web.data:
            for cve in web.data["cves"]:
                try:
                    cvss = float(cve.get("cvss", 0))
                    if cvss >= 9.0: cve_stats["critical"] += 1
                    elif cvss >= 7.0: cve_stats["high"] += 1
                    elif cvss >= 4.0: cve_stats["medium"] += 1
                    else: cve_stats["low"] += 1
                except (ValueError, TypeError) as e: 
                    logger.debug(f"Failed to parse CVSS from '{cve}': {e}")

        # TLDs
        tld_free = tld_scan.data.get("free_count", 0) if tld_scan and tld_scan.data else 0
        tld_diff = tld_scan.data.get("registered_count_diff_owner", 0) if tld_scan and tld_scan.data else 0

        from yads.core.report_translations import get_t
        tr = get_t(lang)
        _yes = tr["label_yes"]
        _no = tr["label_no"]

        row = {
            "ID": t.id,
            tr["col_domain"]: t.domain,
            tr["col_status"]: is_online,
            "Probe": "Active" if (port_scan and port_scan.data and port_scan.data.get("is_active")) else ("Inactive" if port_scan else "-"),
            "HTTP": web.data.get("http_status") if web and web.data else "-",
            "HTTPS": web.data.get("https_status") if web and web.data else "-",
            "HTTPS_Redirect": _yes if (web and web.data and web.data.get("https_redirect")) else _no,
            "Wildcard": _yes if (dns and dns.data and dns.data.get("wildcard_detected")) else _no,
            "Login_Detected": _yes if (web and web.data and web.data.get("is_login_page")) else _no,
            tr["col_ip"]: dns_ip,
            tr["col_subdomains"]: len(dns.data.get("subdomains", [])) if (dns and dns.data) else 0,
            "Scan_Status": t.scan_status,
            tr["col_last_scan"]: results[0].scanned_at.strftime("%Y-%m-%d %H:%M") if results else "-",
            "SSL_Issuer": ssl.data.get("issuer", {}).get("commonName", "") if (ssl and ssl.data and not ssl.data.get("error")) else "",
            tr["col_ssl_expiry"]: ssl.data.get("notAfter", "") if (ssl and ssl.data and not ssl.data.get("error")) else "",
            tr["col_compliance"]: compliant,
            tr["col_non_compliance"]: non_compliant,
            "Web_Server": web.data.get("server_header", "") if web and web.data else "",
            "ASN": infra.data.get("asn", {}).get("asn", "") if infra and infra.data else "",
            "ISP": infra.data.get("asn", {}).get("asn_description", "") if infra and infra.data else "",
            "Secrets_Count": len(web.data.get("secrets", [])) if (web and web.data) else 0,
            tr["col_critical"]: cve_stats["critical"],
            tr["col_high"]: cve_stats["high"],
            tr["col_medium"]: cve_stats["medium"],
            "Takeover_Risks": len(dns.data.get("takeover_risks", [])) if (dns and dns.data) else 0,
            "TLD_Free": tld_free,
            "TLD_Suspect": tld_diff,
            "Created_At": t.created_at.strftime("%Y-%m-%d %H:%M:%S")
        }
        data.append(row)

    df = pd.DataFrame(data)

    output = BytesIO()
    sheet_name = "Ziele" if lang.lower().startswith("de") else "Targets"
    fname = f"yads_targets_export_{lang}.xlsx"
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    output.seek(0)

    headers = {
        'Content-Disposition': f'attachment; filename="{fname}"'
    }
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@router.get("/targets/{target_id}/export/excel")
async def export_target_excel(
    target_id: int,
    lang: str = "en",
    session: Session = Depends(get_session),
    user: User = Depends(RoleChecker(["admin", "tenant_admin", "scanner", "auditor"]))
):
    """Multi-sheet Excel export for a single target (one sheet per scanner module)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    target = session.get(Target, target_id)
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    if user.role != "admin" and target.tenant_id != user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    results = session.exec(
        select(ScanResult).where(ScanResult.target_id == target_id)
        .order_by(ScanResult.scanned_at.desc())
    ).all()
    by_module = {}
    for r in results:
        if r.module_name not in by_module:
            by_module[r.module_name] = r.data or {}

    de = lang.lower().startswith("de")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="EFF4FB")

    def make_header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.freeze_panes = ws["A2"]

    def autofit(ws):
        for col in ws.columns:
            width = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 4, 12), 60)

    def stripe(ws):
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = ALT_FILL

    def kv_sheet(name, pairs):
        ws2 = wb.create_sheet(name)
        make_header(ws2, ["Feld" if de else "Field", "Wert" if de else "Value"])
        for k, v in pairs:
            ws2.append([k, str(v) if v is not None else "-"])
        stripe(ws2)
        autofit(ws2)

    def table_sheet(name, cols, rows_data):
        ws2 = wb.create_sheet(name)
        make_header(ws2, cols)
        for row in rows_data:
            ws2.append(row)
        stripe(ws2)
        autofit(ws2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Übersicht" if de else "Summary"
    make_header(ws, ["Feld" if de else "Field", "Wert" if de else "Value"])
    last_scan = results[0].scanned_at.strftime("%Y-%m-%d %H:%M UTC") if results else "-"
    for row in [
        ("Domain", target.domain),
        ("ID", target.id),
        ("Letzter Scan" if de else "Last Scan", last_scan),
        ("Module", ", ".join(sorted(by_module.keys()))),
        ("Archiviert" if de else "Archived", "Ja" if target.is_archived else "Nein" if de else ("Yes" if target.is_archived else "No")),
    ]:
        ws.append(row)
    autofit(ws)

    d = by_module.get("dns_scanner") or by_module.get("subdomain_scanner") or {}
    if d:
        kv_sheet("DNS", [
            ("Subdomains", len(d.get("subdomains", []))),
            ("Wildcard", d.get("wildcard_detected", "-")),
            ("A Records", ", ".join(d.get("records", {}).get("A", []))),
            ("MX Records", ", ".join(d.get("records", {}).get("MX", []))),
            ("NS Records", ", ".join(d.get("records", {}).get("NS", []))),
            ("TXT Records", " | ".join(d.get("records", {}).get("TXT", []))),
        ])
        subs = d.get("subdomains", [])
        if subs:
            table_sheet("DNS Subdomains", ["Subdomain", "IPs", "Source" if not de else "Quelle"],
                [[s.get("subdomain",""), ", ".join(s.get("ips",[])), s.get("source","")] for s in subs[:500]])

    d = by_module.get("ssl_scanner", {})
    if d:
        cert = d.get("certificate", {})
        kv_sheet("SSL/TLS", [
            ("Issuer CN", cert.get("issuer", {}).get("commonName", d.get("issuer", "-"))),
            ("Subject CN", cert.get("subject", {}).get("commonName", "-")),
            ("Gültig ab" if de else "Valid From", d.get("notBefore", cert.get("notBefore", "-"))),
            ("Gültig bis" if de else "Valid Until", d.get("notAfter", cert.get("notAfter", "-"))),
            ("SANs", ", ".join(d.get("san", []))),
            ("Protocol", d.get("protocol", "-")),
            ("Grade", d.get("grade", "-")),
        ])

    d = by_module.get("web_analyzer", {})
    if d:
        kv_sheet("Web-Analyse" if de else "Web Analysis", [
            ("HTTP Status", d.get("http_status", "-")),
            ("HTTPS Status", d.get("https_status", "-")),
            ("Server", d.get("server_header", "-")),
            ("Technologien" if de else "Technologies", ", ".join(t.get("name","") for t in d.get("technologies", []))),
            ("CVEs", len(d.get("cves", []))),
            ("HTTPS Redirect", d.get("https_redirect", "-")),
        ])

    d = by_module.get("email_security", {})
    if d:
        kv_sheet("E-Mail-Sicherheit" if de else "Email Security", [
            ("Score", d.get("score", "-")),
            ("SPF gültig" if de else "SPF valid", d.get("spf", {}).get("valid", "-")),
            ("SPF Record", d.get("spf", {}).get("record", "-")),
            ("DKIM gültig" if de else "DKIM valid", d.get("dkim", {}).get("valid", "-")),
            ("DMARC gültig" if de else "DMARC valid", d.get("dmarc", {}).get("valid", "-")),
            ("DMARC Policy", d.get("dmarc", {}).get("policy", "-")),
            ("BIMI", d.get("bimi", {}).get("found", "-")),
            ("Probleme" if de else "Issues", " | ".join(str(i) for i in d.get("issues", []))),
        ])

    d = by_module.get("http_headers", {})
    if d:
        kv_sheet("HTTP Headers", [
            ("Score", d.get("score", "-")),
            ("Grade", d.get("grade", "-")),
            ("Fehlend" if de else "Missing", ", ".join(d.get("missing_headers", []))),
            ("Leaky", ", ".join(d.get("leaky_headers", []))),
        ])
        findings = d.get("findings", [])
        if findings:
            table_sheet("HTTP Headers Befunde" if de else "HTTP Headers Findings",
                ["Header", "Schwere" if de else "Severity", "Befund" if de else "Finding"],
                [[f.get("header",""), f.get("severity",""), f.get("message","")] for f in findings])

    d = by_module.get("cookie_scanner", {})
    if d:
        kv_sheet("Cookies", [
            ("Score", d.get("score", "-")),
            ("Anzahl" if de else "Total", len(d.get("cookies", []))),
            ("Unsicher" if de else "Insecure", d.get("insecure_count", "-")),
        ])
        cookies = d.get("cookies", [])
        if cookies:
            table_sheet("Cookies Detail",
                ["Cookie", "Secure", "HttpOnly", "SameSite", "Probleme" if de else "Issues"],
                [[c.get("name",""), c.get("secure","-"), c.get("httponly","-"),
                  c.get("samesite","-"), " | ".join(c.get("issues",[]))] for c in cookies])

    d = by_module.get("cors_scanner", {})
    if d:
        kv_sheet("CORS", [
            ("Score", d.get("score", "-")),
            ("Anfällig" if de else "Vulnerable", d.get("vulnerable", "-")),
            ("Befunde" if de else "Findings", len(d.get("findings", []))),
        ])
        if d.get("findings"):
            table_sheet("CORS Befunde" if de else "CORS Findings",
                ["Origin", "Schwere" if de else "Severity", "Problem" if de else "Issue"],
                [[f.get("origin",""), f.get("severity",""), f.get("message","")] for f in d["findings"]])

    d = by_module.get("subdomain_takeover_scanner", {})
    if d:
        vulns = d.get("vulnerable", [])
        kv_sheet("Subdomain Takeover", [
            ("Score", d.get("score", "-")),
            ("Anfällig" if de else "Vulnerable", len(vulns)),
            ("Geprüft" if de else "Checked", d.get("checked_count", "-")),
        ])
        if vulns:
            table_sheet("Takeover Vulnerable",
                ["Subdomain", "Provider", "CNAME", "Sicherheit" if de else "Confidence"],
                [[v.get("subdomain",""), v.get("provider",""), v.get("cname",""), v.get("confidence","")] for v in vulns])

    d = by_module.get("threat_intel_scanner", {})
    if d:
        kv_sheet("Threat Intel", [
            ("Bedrohungsstufe" if de else "Threat Level", d.get("threat_level", "-")),
            ("AbuseIPDB Score", d.get("abuseipdb", {}).get("abuse_confidence_score", "-")),
            ("OTX Pulses", d.get("otx", {}).get("pulse_count", "-")),
            ("VirusTotal Positives", d.get("virustotal", {}).get("positives", "-")),
        ])

    d = by_module.get("git_exposure_scanner", {})
    if d:
        exposed = d.get("exposed_paths", [])
        kv_sheet("Git-Exposition" if de else "Git Exposure", [
            ("Score", d.get("score", "-")),
            ("Exponiert" if de else "Exposed Paths", len(exposed)),
        ])
        if exposed:
            table_sheet("Git Pfade" if de else "Git Paths",
                ["Pfad" if de else "Path", "Status", "Größe" if de else "Size"],
                [[p.get("path",""), p.get("status",""), p.get("size","")] for p in exposed])

    d = by_module.get("js_secrets_scanner", {})
    if d:
        secrets = d.get("secrets", [])
        kv_sheet("JS-Geheimnisse" if de else "JS Secrets", [
            ("Score", d.get("score", "-")),
            ("Gefunden" if de else "Found", len(secrets)),
            ("Gescannte Dateien" if de else "Files Scanned", d.get("files_scanned", "-")),
        ])
        if secrets:
            table_sheet("JS Secrets Detail",
                ["Typ" if de else "Type", "Datei" if de else "File", "Treffer" if de else "Match"],
                [[s.get("type",""), s.get("file",""), s.get("match","")[:80]] for s in secrets[:200]])

    d = by_module.get("infrastructure_scanner", {})
    if d:
        asn = d.get("asn", {})
        kv_sheet("Infrastruktur" if de else "Infrastructure", [
            ("IP", d.get("ip", "-")),
            ("ASN", asn.get("asn", "-")),
            ("ISP/Org", asn.get("asn_description", "-")),
            ("Land" if de else "Country", asn.get("country", "-")),
            ("Hosting", d.get("hosting_provider", "-")),
            ("CDN", d.get("cdn_provider", "-")),
            ("WAF", d.get("waf", "-")),
        ])

    d = by_module.get("nuclei_scanner", {})
    if d:
        vulns = d.get("findings", d.get("vulnerabilities", []))
        kv_sheet("Nuclei-Schwachstellen" if de else "Nuclei Vulnerabilities", [
            ("Gesamt" if de else "Total", len(vulns)),
            ("Critical", sum(1 for v in vulns if v.get("severity","").lower() == "critical")),
            ("High", sum(1 for v in vulns if v.get("severity","").lower() == "high")),
            ("Medium", sum(1 for v in vulns if v.get("severity","").lower() == "medium")),
        ])
        if vulns:
            table_sheet("Nuclei Befunde" if de else "Nuclei Findings",
                ["ID", "Name", "Schwere" if de else "Severity", "URL"],
                [[v.get("template_id",""), v.get("name",""), v.get("severity",""), v.get("matched_at","")] for v in vulns[:500]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"yads_{target.domain.replace('.','_')}_report_{lang}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get("/targets/{target_id}/export")
async def export_target_pdf(target_id: int, lang: str = "en", session: Session = Depends(get_session), user: User = Depends(RoleChecker(["admin", "scanner"]))):
    """
    Generates a COMPREHENSIVE PDF report for a single target using FPDF.
    Includes full details from all scan modules.
    Supports lang=en (default) and lang=de for German output.
    """
    from fpdf import FPDF
    from io import BytesIO

    target = session.get(Target, target_id)
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    if user.role != "admin" and target.tenant_id != user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    results = session.exec(select(ScanResult).where(ScanResult.target_id == target_id).order_by(ScanResult.scanned_at.desc())).all()

    from yads.core.report_translations import get_t
    tr = get_t(lang)

    # Extract Data
    dns = next((r for r in results if r.module_name == 'dns_scanner'), None)
    web = next((r for r in results if r.module_name == 'web_analyzer'), None)
    ssl = next((r for r in results if r.module_name == 'ssl_scanner'), None)
    infra = next((r for r in results if r.module_name == 'infrastructure_scanner'), None)
    typosquat = next((r for r in results if r.module_name == 'typosquat_scanner'), None)
    visual = next((r for r in results if r.module_name == 'visual_osint'), None)

    _report_title = f'{tr["label_security_report"]}: {target.domain}'
    _page_label = tr["label_page"]
    _confidential = tr["label_confidential"]

    class PDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 15)
            self.cell(0, 10, _report_title, align='C')
            self.ln(12)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'{_page_label} {self.page_no()}', align='C')

        def chapter_title(self, label):
            self.ln(5)
            self.set_font('Helvetica', 'B', 12)
            self.set_fill_color(200, 220, 255)
            self.set_x(self.l_margin)
            self.cell(0, 8, label, fill=True, align='L')
            self.ln(10)

        def content_text(self, text):
            self.set_font('Helvetica', '', 10)
            self.set_x(self.l_margin)
            self.multi_cell(0, 5, text)
            self.ln(1)

        def section_kv(self, key, value):
            # Safe string conversion
            text_value = str(value) if value is not None else "N/A"
            if not text_value: text_value = "N/A"
            
            # Check for page break needed
            if self.get_y() > 270:
                self.add_page()
            
            self.set_x(self.l_margin)
            
            # Key
            self.set_font('Helvetica', 'B', 10)
            self.cell(45, 5, f"{key}:", align='L')
            
            # Value
            self.set_font('Helvetica', '', 10)
            # Calculate remaining width strictly
            # cell moved cursor 45 to the right (implicit)
            # We want to use the rest of the line
            remaining_w = self.w - self.l_margin - self.r_margin - 45
            self.multi_cell(remaining_w, 5, text_value, align='L')
            
        def section_header(self, title):
             self.ln(3)
             if self.get_y() > 270: self.add_page()
             self.set_x(self.l_margin)
             self.set_font('Helvetica', 'B', 10)
             self.cell(0, 6, title, align='L')
             self.ln(6)
            
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # helper for checking data
    def safe_get(obj, *keys, default="N/A"):
        try:
            val = obj
            for k in keys:
                val = val.get(k, {})
            return val if val else default
        except Exception as e:
            logger.debug(f"Error getting keys: {e}")
            return default

    _titles = {
        "overview": {"en": "Target Overview", "de": "Zielübersicht"},
        "dns": {"en": "DNS Analysis", "de": "DNS-Analyse"},
        "web": {"en": "Web Analysis", "de": "Web-Analyse"},
        "ssl": {"en": "SSL Configuration", "de": "SSL-Konfiguration"},
        "infra": {"en": "Infrastructure", "de": "Infrastruktur"},
        "typosquat": {"en": "Typosquatting", "de": "Typosquatting"},
        "visual": {"en": "Visual OSINT", "de": "Visuelles OSINT"},
        "findings": {"en": "Security Findings", "de": "Sicherheitsbefunde"},
    }
    _lang_key = "de" if lang.lower().startswith("de") else "en"

    def _title(key):
        return _titles.get(key, {}).get(_lang_key, key)

    # 1. Overview
    pdf.chapter_title(_title("overview"))
    pdf.section_kv("Domain", target.domain)
    pdf.section_kv("Target ID", target.id)
    pdf.section_kv("Scan Status", target.scan_status)
    pdf.section_kv("Scan Progress", f"{target.scan_progress}%" if target.scan_progress else "N/A")
    pdf.section_kv("Created At", target.created_at.strftime("%Y-%m-%d %H:%M:%S") if target.created_at else "N/A")

    # 2. DNS Analysis
    pdf.chapter_title(_title("dns"))
    if dns and dns.data:
        ip_records = dns.data.get("a_records", [])
        mx_records = dns.data.get("mx_records", [])
        txt_records = dns.data.get("txt_records", [])
        cname_records = dns.data.get("cname_records", [])
        ns_records = dns.data.get("ns_records", [])
        soa_records = dns.data.get("soa_records", [])
        subdomains = dns.data.get("subdomains", [])
        
        pdf.section_kv("A Records", ", ".join(ip_records) if ip_records else "None")
        pdf.section_kv("MX Records", ", ".join(mx_records) if mx_records else "None")
        pdf.section_kv("NS Records", ", ".join(ns_records) if ns_records else "None")
        pdf.section_kv("CNAME Records", ", ".join(cname_records) if cname_records else "None")
        pdf.section_kv("SOA Records", ", ".join(soa_records) if soa_records else "None")
        
        if txt_records:
             pdf.section_header("TXT Records")
             pdf.set_font('Courier', '', 8)
             for txt in txt_records:
                 if pdf.get_y() > 270: pdf.add_page()
                 pdf.set_x(pdf.l_margin)
                 pdf.multi_cell(0, 4, txt)
                 pdf.ln(1)

        if subdomains:
            pdf.section_header(f"Subdomains Found ({len(subdomains)})")
            pdf.set_font('Helvetica', '', 9)
            for s in subdomains:
                if pdf.get_y() > 270: pdf.add_page()
                line = f"- {s.get('subdomain')} ({', '.join(s.get('ips', []))})"
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(0, 5, line)
    else:
        pdf.content_text("No DNS data available.")
    
    # 3. Web Analysis
    pdf.chapter_title(_title("web"))
    if web and web.data:
        pdf.section_kv("Title", web.data.get("title", "No Title"))
        pdf.section_kv("Server", web.data.get("server_header", "Unknown"))
        pdf.section_kv("Status Code", web.data.get("status_code", "Unknown"))
        
        tech = web.data.get("technologies", [])
        pdf.section_kv("Technologies", ", ".join(tech) if tech else "None detected")
        
        headers_dict = web.data.get("http_headers", {})
        if headers_dict:
            pdf.section_header("HTTP Headers")
            pdf.set_font('Courier', '', 8)
            for k, v in headers_dict.items():
                if pdf.get_y() > 270: pdf.add_page()
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(0, 4, f"{k}: {v}")
                
        redirects = web.data.get("redirect_chain", [])
        if redirects:
            pdf.section_header("Redirect Chain")
            pdf.set_font('Helvetica', '', 9)
            for r in redirects:
                 if pdf.get_y() > 270: pdf.add_page()
                 pdf.set_x(pdf.l_margin)
                 pdf.cell(0, 5, f"-> {r}", ln=True)
                 
        risk = web.data.get("risk_hints", [])
        if risk:
            pdf.section_header("Risk Indicators")
            pdf.set_text_color(200, 50, 50)
            pdf.set_font('Helvetica', 'B', 9)
            for r in risk:
                if pdf.get_y() > 270: pdf.add_page()
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(0, 5, f"[!] {r}")
            pdf.set_text_color(0, 0, 0)
            
    else:
        pdf.content_text("No Web analysis data available.")
    
    # 4. SSL Configuration
    pdf.chapter_title(_title("ssl"))
    if ssl and ssl.data:
        if ssl.data.get("error"):
            pdf.set_text_color(200, 50, 50)
            pdf.content_text(f"Error: {ssl.data.get('error')}")
            pdf.set_text_color(0, 0, 0)
        else:
            issuer = safe_get(ssl.data, "issuer", "commonName")
            org = safe_get(ssl.data, "issuer", "organizationName")
            
            subject_cn = safe_get(ssl.data, "subject", "commonName")
            
            pdf.section_kv("Issued To", subject_cn)
            pdf.section_kv("Issued By", f"{issuer} ({org})")
            pdf.section_kv("Valid From", ssl.data.get("notBefore"))
            pdf.section_kv("Valid To", ssl.data.get("notAfter"))
            
            sans = ssl.data.get("subjectAltName", [])
            if sans:
                 pdf.section_header(f"Subject Alternative Names ({len(sans)})")
                 pdf.set_font('Helvetica', '', 8)
                 # Join them into a block of text
                 sans_text = ", ".join([s[1] for s in sans])
                 pdf.set_x(pdf.l_margin)
                 pdf.multi_cell(0, 4, sans_text)

    else:
        pdf.content_text("No SSL data available.")

    # 5. Infrastructure
    pdf.chapter_title(_title("infra"))
    if infra and infra.data:
        asn = safe_get(infra.data, "asn", "asn")
        org = safe_get(infra.data, "asn", "asn_description")
        country = safe_get(infra.data, "geoip", "country_name")
        cloud = infra.data.get("cloud_provider", "Unknown")
        
        pdf.section_kv("IP Address", infra.data.get("ip"))
        pdf.section_kv("ASN", f"{asn} ({org})")
        pdf.section_kv("Location", country)
        pdf.section_kv("Cloud Provider", cloud)
        
        buckets = infra.data.get("buckets", [])
        if buckets:
             pdf.section_header("Storage Buckets")
             pdf.set_font('Helvetica', '', 9)
             for b in buckets:
                 if pdf.get_y() > 270: pdf.add_page()
                 status = b.get('status')
                 pdf.set_x(pdf.l_margin)
                 pdf.multi_cell(0, 5, f"- {b.get('url')} [{status}]")

    else:
        pdf.content_text("No Infrastructure data available.")

    # 6. Typosquatting
    pdf.chapter_title(_title("typosquat"))
    if typosquat and typosquat.data:
        found = typosquat.data.get("found", [])
        scanned = typosquat.data.get("scanned_count", 0)
        total = typosquat.data.get("total_variations", 0)
        
        pdf.section_kv("Variations Checked", f"{scanned} / {total}")
        pdf.section_kv("Suspicious Found", len(found))
        
        if found:
            pdf.ln(2)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 5, "Detected Squats:", new_x="LMARGIN", new_y="NEXT")
            
            # Simple table for squats
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(60, 6, "Domain", border=1)
            pdf.cell(40, 6, "IP", border=1)
            pdf.cell(60, 6, "Registrar/Info", border=1)
            pdf.ln()
            
            pdf.set_font('Helvetica', '', 8)
            for sq in found:
                domain = str(sq.get('domain', ''))[:35]
                ip = str(sq.get('ip', ''))[:20]
                info = str(sq.get('fuzzer', ''))[:35]
                
                pdf.cell(60, 6, domain, border=1)
                pdf.cell(40, 6, ip, border=1)
                pdf.cell(60, 6, info, border=1)
                pdf.ln()
    else:
        pdf.chapter_body("No Typosquatting data available.")

    # 7. Visual OSINT
    pdf.chapter_title(_title("visual"))
    if visual and visual.data:
        logos = visual.data.get("logos", [])
        if logos:
             pdf.section_kv("Logos Found", len(logos))
             pdf.ln()
             for logo in logos:
                 pdf.set_font('Helvetica', '', 9)
                 pdf.multi_cell(0, 5, f" - {logo.get('source')} ({logo.get('type')}): {logo.get('url')}")
        else:
             pdf.chapter_body("No external visual identity found.")
    else:
        pdf.chapter_body("No Visual OSINT data available.")

    # Output
    output_pdf = BytesIO(pdf.output()) 
    
    headers = {
        'Content-Disposition': f'attachment; filename="target_{target.domain}_report.pdf"'
    }
    return StreamingResponse(output_pdf, headers=headers, media_type='application/pdf')
